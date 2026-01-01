import os
import re
from pathlib import Path
from supabase import create_client
import docx
import pypdf
import openpyxl

# Configuration
DOCS_DIR = Path("pdf and docs")
MIN_CHUNK_WORDS = 100
MAX_CHUNK_WORDS = 500
OVERLAP_WORDS = 50

# Target files to process (Exact filenames)
TARGET_FILES = [
    "Mathematics - I (New Syllabus) (1).docx",
    "BS DSAI_Philosophy and Sociology.docx",
    "Course Details_Prof. Pooja Jain.docx",
    "Course_Outline DS&AI - Statistics-Final.docx",
    "1_Course_Outline_Positive Psychology_DSAI.docx",
    "SCHEDULE _ DSAI _ BATCH 2025-29 _ SEM-I.xlsx",
    "IIM_Sambalpur_Data_Science_Programme_calender.pdf"
]

def get_supabase_client():
    url = os.getenv('NEXT_PUBLIC_SUPABASE_URL')
    key = os.getenv('SUPABASE_SERVICE_ROLE_KEY')
    if not url or not key:
        raise ValueError("Missing Supabase credentials")
    return create_client(url, key)

import zipfile
import xml.etree.ElementTree as ET

def extract_text_from_docx(filepath):
    """
    Extract text from DOCX file using XML parsing.
    This is more robust for complex layouts/textboxes than python-docx.
    """
    try:
        with zipfile.ZipFile(filepath) as z:
            xml_content = z.read('word/document.xml')
            
        tree = ET.fromstring(xml_content)
        # Namespace map often needed for w:t
        namespaces = {'w': 'http://schemas.openxmlformats.org/wordprocessingml/2006/main'}
        
        text_parts = []
        for node in tree.iter():
            if node.tag.endswith('}t'): # w:t tags contain text
                if node.text:
                    text_parts.append(node.text)
            elif node.tag.endswith('}p'): # w:p is paragraph, add newline
                text_parts.append('\n')
                
        # Join and clean up
        full_text = "".join(text_parts)
        # Collapse excessive newlines
        return re.sub(r'\n\s*\n', '\n', full_text).strip()
        
    except Exception as e:
        print(f"Error reading DOCX (XML mode) {filepath}: {e}")
        return ""

def extract_text_from_pdf(filepath):
    """Extract text from PDF file."""
    try:
        text = []
        with open(filepath, 'rb') as f:
            reader = pypdf.PdfReader(f)
            for page in reader.pages:
                page_text = page.extract_text()
                if page_text:
                    text.append(page_text.strip())
        return "\n".join(text)
    except Exception as e:
        print(f"Error reading PDF {filepath}: {e}")
        return ""

def extract_text_from_xlsx(filepath):
    """Extract text from Excel file."""
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        text = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            text.append(f"=== Sheet: {sheet} ===")
            for row in ws.iter_rows(values_only=True):
                # Filter None values and join
                row_text = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
                if row_text:
                    text.append(" | ".join(row_text))
        return "\n".join(text)
    except Exception as e:
        print(f"Error reading XLSX {filepath}: {e}")
        return ""

def extract_tags(text, filename):
    """Extract relevant tags from text and filename."""
    text_lower = text.lower() + " " + filename.lower()
    tags = []
    
    tag_keywords = {
        'mathematics': ['mathematics', 'calculus', 'algebra', 'statistics', 'probability'],
        'data_science': ['data science', 'machine learning', 'deep learning', 'ai', 'artificial intelligence', 'dsai'],
        'syllabus': ['syllabus', 'course outline', 'curriculum', 'topics'],
        'professor': ['professor', 'prof', 'faculty'],
        'schedule': ['schedule', 'timetable', 'calendar', 'dates'],
        'yoga': ['yoga', 'meditation'],
        'psychology': ['psychology', 'behavioral'],
        'philosophy': ['philosophy', 'sociology', 'ethics']
    }
    
    for tag, keywords in tag_keywords.items():
        if any(kw in text_lower for kw in keywords):
            tags.append(tag)
            
    return list(set(tags))

def chunk_text(text, source_url, page_title):
    """Chunk text into segments."""
    words = text.split()
    chunks = []
    start = 0
    chunk_num = 0
    
    while start < len(words):
        end = min(start + MAX_CHUNK_WORDS, len(words))
        chunk_words = words[start:end]
        chunk_text = ' '.join(chunk_words)
        
        if len(chunk_words) >= MIN_CHUNK_WORDS or start + len(chunk_words) >= len(words):
            chunks.append({
                'source_url': source_url,
                'page_title': f"{page_title} (Chunk {chunk_num + 1})",
                'content': chunk_text,
                'word_count': len(chunk_words),
                'tags': extract_tags(chunk_text, page_title)
            })
            chunk_num += 1
        
        start = end - OVERLAP_WORDS if end < len(words) else len(words)
    
    return chunks

def main():
    print("=" * 60)
    print("Re-uploading Specific Documents to Supabase")
    print("=" * 60)
    
    supabase = get_supabase_client()
    
    for filename in TARGET_FILES:
        filepath = DOCS_DIR / filename
        if not filepath.exists():
            print(f"⚠️ File not found: {filename}")
            continue
            
        print(f"\nProcessing: {filename}")
        
        # 1. Delete existing chunks for this file
        # We search by partial source_url match (filename)
        try:
            print(f"  🗑️ Deleting old chunks for {filename}...")
            # Note: This is an approximation. Ideally we'd store a file_id. 
            # But ilike on source_url should cover 'local_file_Upload/Filename' patterns if we used them, 
            # or simply verify against 'title'. based on previous script 'page_title' or 'source_url'
            # Previous script used `doc['page_title']` which was filename based.
            
            # Let's try to delete by page_title ilike
            supabase.table('chunks').delete().ilike('page_title', f'%{filename}%').execute()
            print("  ✅ Old chunks deleted (if any)")
        except Exception as e:
            print(f"  ❌ Error deleting old chunks: {e}")
            
        # 2. Extract Text
        content = ""
        if filepath.suffix.lower() == '.docx':
            content = extract_text_from_docx(filepath)
        elif filepath.suffix.lower() == '.pdf':
            content = extract_text_from_pdf(filepath)
        elif filepath.suffix.lower() == '.xlsx':
            content = extract_text_from_xlsx(filepath)
            
        if not content:
            print("  ❌ No content extracted. Skipping.")
            continue
            
        print(f"  📄 Extracted {len(content)} characters")
        
        # 3. Chunk
        # Use pseudo-URL for local files
        source_url = f"file://{filename}"
        chunks = chunk_text(content, source_url, filename)
        print(f"  🧩 Created {len(chunks)} chunks")
        
        # 4. Upload
        if not chunks:
            continue
            
        records = []
        for chunk in chunks:
            # Sanitize
            safe_content = chunk['content'].encode('ascii', 'ignore').decode('ascii')
            records.append({
                'source_url': chunk['source_url'],
                'page_title': chunk['page_title'],
                'content': safe_content,
                'word_count': chunk['word_count'],
                'tags': chunk['tags'],
                'embedding': [0.0] * 384 
            })
            
        try:
            supabase.table('chunks').insert(records).execute()
            print(f"  🚀 Uploaded {len(records)} chunks successfully")
        except Exception as e:
            print(f"  ❌ Upload failed: {e}")

if __name__ == "__main__":
    main()
